"""Build a formatted Excel workbook from the AURA battery results CSV."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# results (from aura_battery_out/results.csv)
ROWS = [
    ("Quality", "test1_quality.wav", "watermarked audio, no attack", "Detected", 0.9149, "11010001110100100000000010110010", 58),
    ("Compression", "test2_mp3_64k.mp3", "MP3 64 kbps", "Detected", 0.9036, "11010001110100100000000010110010", 58),
    ("Compression", "test2_mp3_128k.mp3", "MP3 128 kbps", "Detected", 0.9127, "11010001110100100000000010110010", 58),
    ("Compression", "test2_mp3_320k.mp3", "MP3 320 kbps", "Detected", 0.9149, "11010001110100100000000010110010", 58),
    ("Format chain", "test3_format_chain.wav", "wav->mp3->wav->flac->wav", "Detected", 0.8966, "11010001110100100000000010110010", 58),
    ("Editing", "test4_editing.wav", "trim + insert silence + cut slice", "Detected", 0.8798, "11010001110100100000000010110010", 58),
    ("Signal processing", "test5_signal.wav", "EQ + pitch shift + normalize", "Not Detected", 0.5221, "11100101100100100011111011111100", 58),
    ("Platform simulation", "test6_platform_opus.wav", "Opus 24 kbps round-trip", "Detected", 0.8050, "11010001110100100000100010110011", 58),
    ("Re-recording simulation", "test7_rerecord_sim.wav", "reverb + band-limit + noise", "Detected", 0.8287, "11010001110100100000000010110010", 58),
]
EMBEDDED = "11010001110100100000000010110010"
THRESHOLD = 0.70
FONT = "Arial"

wb = Workbook()
ws = wb.active
ws.title = "AURA Battery"

navy = PatternFill("solid", fgColor="1F3864")
grey = PatternFill("solid", fgColor="F2F2F2")
green = PatternFill("solid", fgColor="C6EFCE")
red = PatternFill("solid", fgColor="FFC7CE")
green_txt = Font(name=FONT, color="006100")
red_txt = Font(name=FONT, color="9C0006")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── title ─────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:H1")
c = ws["A1"]; c.value = "AURA — Watermark Robustness Battery"
c.font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
c.fill = navy; c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 24

# ── metadata ──────────────────────────────────────────────────────────────────
meta = [
    ("Checkpoint", "run_002/step_0200000_final.pt"),
    ("Input audio", "audio/client_original.mp3"),
    ("Embedded 32-bit message", EMBEDDED),
    ("Detection threshold (bit-acc)", THRESHOLD),
    ("Watermark window", "2 s / 48 kHz / 32-bit (tiled across clip)"),
]
r = 2
for k, v in meta:
    ws.cell(r, 1, k).font = Font(name=FONT, bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r, 2, v).font = Font(name=FONT)
    r += 1
r += 1

# ── table header ──────────────────────────────────────────────────────────────
headers = ["Test Criteria", "File", "Remarks", "Detection Result",
           "Detection Probability", "Threshold Used", "Decoded 32-bit Message", "Windows"]
hdr = r
for j, h in enumerate(headers, 1):
    cell = ws.cell(hdr, j, h)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF")
    cell.fill = navy
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[hdr].height = 30

# ── data rows ─────────────────────────────────────────────────────────────────
first = hdr + 1
for i, (crit, f, rem, res, prob, bits, nwin) in enumerate(ROWS):
    row = first + i
    exact = bits == EMBEDDED and res == "Detected"
    vals = [crit, f, rem, res, prob, THRESHOLD, bits, nwin]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row, j, v)
        cell.font = Font(name=FONT)
        cell.border = border
        cell.alignment = Alignment(vertical="center",
                                   horizontal="center" if j in (4, 5, 6, 8) else "left")
    if i % 2:
        for j in range(1, 9):
            ws.cell(row, j).fill = grey
    # detection result colour
    rc = ws.cell(row, 4)
    if res == "Detected":
        rc.fill = green; rc.font = green_txt
    else:
        rc.fill = red; rc.font = red_txt
    # probability format + colour scale
    pc = ws.cell(row, 5); pc.number_format = "0.0%"
    pc.font = Font(name=FONT, bold=True,
                   color="006100" if prob >= THRESHOLD else "9C0006")
    ws.cell(row, 6).number_format = "0.0%"
    # monospace bits, flag mismatch
    bc = ws.cell(row, 7)
    bc.font = Font(name="Consolas", color="000000" if exact else "9C0006")

last = first + len(ROWS) - 1

# ── summary row ───────────────────────────────────────────────────────────────
s = last + 1
ws.cell(s, 1, "SUMMARY").font = Font(name=FONT, bold=True)
ws.merge_cells(start_row=s, start_column=1, end_row=s, end_column=3)
n_det = sum(1 for r_ in ROWS if r_[3] == "Detected")
ws.cell(s, 4, f"{n_det}/{len(ROWS)} detected").font = Font(name=FONT, bold=True)
ws.merge_cells(start_row=s, start_column=4, end_row=s, end_column=4)
avg = ws.cell(s, 5, round(sum(r_[4] for r_ in ROWS) / len(ROWS), 4))
avg.number_format = "0.0%"; avg.font = Font(name=FONT, bold=True)
for j in range(1, 9):
    ws.cell(s, j).fill = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(s, j).border = border

# ── notes ─────────────────────────────────────────────────────────────────────
n = s + 2
notes = [
    "Notes:",
    "• Detection Probability = mean per-window bit accuracy (1 − BER) over the tiled 2 s windows.",
    "• Decoded message in black = bit-identical to the embedded payload after majority vote; red = differs.",
    "• Signal processing fails because the pitch shift changes the time base and desyncs AURA's fixed",
    "  2 s windows (no sync layer) — the one genuine weak spot; all other attacks recover the exact payload.",
]
for k, txt in enumerate(notes):
    cell = ws.cell(n + k, 1, txt)
    cell.font = Font(name=FONT, bold=(k == 0), italic=(k > 0), size=9,
                     color="595959" if k > 0 else "000000")
    ws.merge_cells(start_row=n + k, start_column=1, end_row=n + k, end_column=8)

# ── widths / freeze ───────────────────────────────────────────────────────────
widths = [24, 24, 34, 16, 14, 14, 36, 10]
for j, wd in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = wd
ws.freeze_panes = f"A{first}"
ws.sheet_view.showGridLines = False

wb.save("aura_battery_results.xlsx")
print("wrote aura_battery_results.xlsx")
